#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
露台石英砖铺设布局Excel生成器
生成专业的Excel文件，包含完整的铺设布局和材料计算
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

def create_layout_excel():
    """创建完整的铺设布局Excel文件"""
    
    # 创建工作簿
    wb = Workbook()
    
    # 创建工作表
    ws_layout = wb.active
    ws_layout.title = "铺设布局图"
    ws_calc = wb.create_sheet("材料计算")
    
    # 设置样式
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    support_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色支撑器
    tile_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")    # 浅蓝色石英砖
    calc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # 浅黄色计算区
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    normal_font = Font(size=10)
    title_font = Font(bold=True, size=14)
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # === 铺设布局图工作表 ===
    
    # 标题
    ws_layout["A1"] = "露台石英砖铺设布局图"
    ws_layout["A1"].font = title_font
    ws_layout["A1"].alignment = center_alignment
    ws_layout["A1"].fill = header_fill
    ws_layout.merge_cells("A1:U1")
    
    # 项目信息
    info_start_row = 3
    info_data = [
        ["项目信息", ""],
        ["铺设面积", "73.6平方米"],
        ["石英砖规格", "600mm × 600mm"],
        ["支撑器间距", "600mm × 600mm"],
        ["", ""],
        ["图例说明", ""],
        ["●", "支撑器位置"],
        ["砖", "石英砖区域"],
        ["", ""],
    ]
    
    for i, (key, value) in enumerate(info_data):
        row = info_start_row + i
        ws_layout[f"A{row}"] = key
        ws_layout[f"B{row}"] = value
        if key == "项目信息" or key == "图例说明":
            ws_layout[f"A{row}"].font = Font(bold=True, size=11)
        if key == "●":
            ws_layout[f"A{row}"].fill = support_fill
        elif key == "砖":
            ws_layout[f"A{row}"].fill = tile_fill
    
    # 布局网格开始行
    grid_start_row = 15
    
    # 列标题 (A-U)
    for col_idx in range(21):  # A到U共21列
        col_letter = get_column_letter(col_idx + 1)
        ws_layout[f"{col_letter}{grid_start_row}"] = col_letter
        ws_layout[f"{col_letter}{grid_start_row}"].font = header_font
        ws_layout[f"{col_letter}{grid_start_row}"].fill = header_fill
        ws_layout[f"{col_letter}{grid_start_row}"].alignment = center_alignment
    
    # 生成布局网格
    tile_counter = 1
    
    for row_idx in range(22):  # 22行网格
        excel_row = grid_start_row + 1 + row_idx
        
        # 行号
        ws_layout[f"A{excel_row}"] = str(row_idx + 1)
        ws_layout[f"A{excel_row}"].font = header_font
        ws_layout[f"A{excel_row}"].fill = header_fill
        ws_layout[f"A{excel_row}"].alignment = center_alignment
        
        for col_idx in range(1, 21):  # B到U列
            col_letter = get_column_letter(col_idx + 1)
            cell = ws_layout[f"{col_letter}{excel_row}"]
            
            # 支撑器位置 (奇数行奇数列，偶数行偶数列)
            if (row_idx % 2 == 0 and col_idx % 2 == 1) or (row_idx % 2 == 1 and col_idx % 2 == 0):
                cell.value = "●"
                cell.fill = support_fill
            # 石英砖区域
            elif row_idx % 2 == 1 and col_idx % 2 == 1:
                cell.value = f"砖{tile_counter}"
                cell.fill = tile_fill
                tile_counter += 1
            
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.font = Font(size=8)
    
    # 设置列宽
    for col_idx in range(1, 22):
        col_letter = get_column_letter(col_idx)
        ws_layout.column_dimensions[col_letter].width = 4
    
    # 设置行高
    for row_idx in range(grid_start_row, grid_start_row + 23):
        ws_layout.row_dimensions[row_idx].height = 20
    
    # === 材料计算工作表 ===
    
    # 标题
    ws_calc["A1"] = "露台石英砖材料计算表"
    ws_calc["A1"].font = title_font
    ws_calc["A1"].alignment = center_alignment
    ws_calc["A1"].fill = header_fill
    ws_calc.merge_cells("A1:F1")
    
    # 项目基本信息
    calc_data = [
        ["", "", "", "", "", ""],
        ["项目基本信息", "", "", "", "", ""],
        ["项目名称", "露台石英砖架空铺设", "", "", "", ""],
        ["铺设面积", "73.6", "平方米", "", "", ""],
        ["石英砖规格", "600×600", "mm", "", "", ""],
        ["铺设方式", "支撑器架空铺设", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["材料计算明细", "", "", "", "", ""],
        ["材料名称", "规格", "理论数量", "调整系数", "最终数量", "单位"],
        ["石英砖", "600×600mm", "204.44", "+5%", "215", "块"],
        ["支撑器", "可调高度", "231", "-10%", "208", "个"],
        ["", "", "", "", "", ""],
        ["计算过程详解", "", "", "", "", ""],
        ["石英砖计算", "", "", "", "", ""],
        ["单块面积", "0.36", "平方米", "", "", ""],
        ["理论数量", "73.6÷0.36", "204.44块", "", "", ""],
        ["损耗率", "5%", "", "", "", ""],
        ["最终需求", "204.44×1.05", "215块", "", "", ""],
        ["", "", "", "", "", ""],
        ["支撑器计算", "", "", "", "", ""],
        ["网格布置", "600mm间距", "", "", "", ""],
        ["横向支撑线", "21", "条", "", "", ""],
        ["纵向支撑线", "11", "条", "", "", ""],
        ["网格交点", "21×11", "231个", "", "", ""],
        ["不规则调整", "-10%", "", "", "", ""],
        ["最终需求", "231×0.9", "208个", "", "", ""],
        ["", "", "", "", "", ""],
        ["施工要点", "", "", "", "", ""],
        ["1. 支撑器按600mm网格精确定位", "", "", "", "", ""],
        ["2. 每块石英砖四角必须有支撑点", "", "", "", "", ""],
        ["3. 边缘每600mm设置支撑器", "", "", "", "", ""],
        ["4. 支撑器高度根据现场调节", "", "", "", "", ""],
        ["5. 石英砖接缝控制在2-3mm", "", "", "", "", ""],
    ]
    
    # 填充数据
    for row_idx, row_data in enumerate(calc_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_calc.cell(row=row_idx, column=col_idx)
            cell.value = value
            
            # 设置样式
            if row_idx == 2 or row_idx == 8 or row_idx == 13 or row_idx == 20 or row_idx == 28:
                cell.font = Font(bold=True, size=11)
            elif row_idx == 9:  # 表头
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            elif row_idx in [10, 11]:  # 计算结果
                cell.fill = calc_fill
            
            cell.border = thin_border
    
    # 设置列宽
    ws_calc.column_dimensions['A'].width = 25
    ws_calc.column_dimensions['B'].width = 15
    ws_calc.column_dimensions['C'].width = 12
    ws_calc.column_dimensions['D'].width = 12
    ws_calc.column_dimensions['E'].width = 12
    ws_calc.column_dimensions['F'].width = 8
    
    # 保存文件
    filename = "露台石英砖铺设布局计算表.xlsx"
    wb.save(filename)
    print(f"Excel文件已生成: {filename}")
    
    return filename

if __name__ == "__main__":
    try:
        create_layout_excel()
        print("✅ Excel文件生成成功！")
        print("📊 文件包含两个工作表：")
        print("   1. 铺设布局图 - 完整的网格布局和图例")
        print("   2. 材料计算 - 详细的计算过程和结果")
        print("🎨 使用颜色标注：")
        print("   - 红色：支撑器位置")
        print("   - 浅蓝色：石英砖区域")
        print("   - 浅黄色：计算结果")
    except ImportError as e:
        print("❌ 缺少依赖库，请安装：")
        print("pip install openpyxl pandas")
        print(f"错误详情：{e}")
    except Exception as e:
        print(f"❌ 生成Excel文件时出错：{e}") 